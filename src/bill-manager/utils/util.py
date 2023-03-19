"""
Author       : UniMars
Date         : 2023-01-13 23:25:17
LastEditors  : UniMars
LastEditTime : 2023-03-18 14:24:45
Description  : file head
"""
import os.path
from os.path import join
from os import listdir
from shutil import move

from openpyxl import load_workbook
from pandas import DataFrame

JRZZLJ = "计入总账逻辑"


def bill_filter(word, ban_list: list[str]) -> bool:
    if not isinstance(word, str):
        return True
    return not word.startswith(tuple(ban_list))


def plus_minus_filter(x):
    if x.startswith("收入"):
        return 1
    elif x.startswith("支出"):
        return -1
    return 0


class BillReader(object):
    __slots__ = '_file', "_type", "_host_name", "_output_path", "_time_filter"

    def __init__(self, file=""):
        self._type = ""
        self._file = file
        # if file != "":
        #     self.read_file(file)

    def read_file(self, filepath=""):
        assert filepath != "" or self._file != "", "no input location"
        filepath = self._file if filepath == "" else filepath
        self._file = filepath
        suffix = filepath.split('.')[-1]
        assert suffix in ["csv", "xlsx", "xls"], "file type is not sheet"
        self._type = suffix
        # print(f"type:{suffix}\n")


class BillWriter(object):
    __slots__ = "_output_path", "_time_filter"

    def __init__(self, output_path: str = r"C:\Users\Lenovo\OneDrive\Documents\个人财务管理.xlsx", time_filter=True):
        self._output_path = output_path
        self._time_filter = time_filter

    def starter(self, path):
        for i in listdir(path):
            if i == "old":
                continue
            yield i
            old_file = join(path, i)
            new_path = join(path, "old")
            if not os.path.exists(new_path):
                os.makedirs(new_path)
            move(src=old_file, dst=new_path)

    def write(self, df: DataFrame):
        book = load_workbook(self._output_path)
        if '流水表' in book.sheetnames:
            sh = book['流水表']
        else:
            sh = book.create_sheet('流水表')
            sh.append(["交易时间",
                       "来源",
                       "收/支",
                       "交易类型",
                       "交易对象",
                       "商品",
                       "金额",
                       "母类别",
                       "子类别",
                       "总类别",
                       "备注",
                       "收支逻辑",
                       "计入总账逻辑",
                       "乘后金额"])
        t = 0
        if df.empty:
            print("\n没有数据写入")
            return

        df = df.loc[df[JRZZLJ] == 1]
        max_row = sh.max_row
        while True:
            last_line = sh[max_row]
            if last_line[0].value and last_line[1].value and str(last_line[0].value).strip() and str(
                    last_line[1].value).strip():
                break
            sh.delete_rows(max_row)
            max_row -= 1

        if self._time_filter:
            t = sh.cell(sh.max_row, 1).value
        if t and t != "交易时间":
            df = df.loc[df['交易时间'] > t]

        for _, row in df.iterrows():
            row_list = list(row)
            sh.append(row_list)
        book.save(self._output_path)
        print("\n写入完成")
